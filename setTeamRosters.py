import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from getRosters import get_rosters

# Function to set team rosters in an Excel file
def set_team_rosters(workbook, excel_filename, sleeperUsername, playerData):
    if "Team Rosters" in workbook.sheetnames:
        worksheet = workbook["Team Rosters"]
    else:
        worksheet = workbook.create_sheet("Team Rosters")

    # Set all columns to a width of 20
    max_columns = 10  # Adjust based on the maximum number of columns used
    for col_idx in range(1, max_columns + 1):
        column_letter = get_column_letter(col_idx)  # Convert index to column letter (A, B, etc.)
        worksheet.column_dimensions[column_letter].width = 20

    # Clear message in a specific cell
    worksheet.cell(row=5, column=5, value="Clearing...")

    # Clear the existing content
    worksheet.delete_rows(1, worksheet.max_row)

    # Set the height for each row in the specified range
    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 21

    # Get the team rosters
    team_rosters = get_rosters(sleeperUsername, playerData)
    current_row = 2

    for roster in team_rosters:
        top_of_roster = current_row

        # Add League details with styles
        font_bold = Font(bold=True)
        fill_light_blue = PatternFill(start_color="DBE8F0", end_color="DBE8F0", fill_type="solid")

        details = [
            "League:",
            "Num Teams:",
            "Quarterbacks:",
            "PPR:",
            "Tight Ends Premium:",
            "Starters:",
            "Record:",
        ]

        # Write the league details
        for i, detail in enumerate(details):
            worksheet.cell(row=current_row + i, column=1, value=detail).font = font_bold
            worksheet.cell(row=current_row + i, column=1).fill = fill_light_blue

        # Write the corresponding data from the roster
        league_details = [
            roster["league_name"],
            roster["total_teams"],
            roster["qb"],
            roster["ppr"],
            roster["tep"],
            roster["starters"],
            roster["record"],
        ]

        for i, value in enumerate(league_details):
            worksheet.cell(row=current_row + i, column=2, value=value).alignment = Alignment(horizontal="left")

        current_row += 7

        # Add position details with specific styles and alignment
        positions = ["Quarterbacks", "Running Backs", "Wide Receivers", "Tight Ends", "Draft Capital"]
        position_colors = [
            "E9BDAF",  # Light orange
            "E3EBA7",  # Light green
            "B7D0E0",  # Light blue
            "F8D79D",  # Light yellow
            "C0C0C0",  # Gray
        ]

        for i, position in enumerate(positions):
            cell = worksheet.cell(row=current_row, column=i + 1, value=position)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.fill = PatternFill(start_color=position_colors[i], end_color=position_colors[i], fill_type="solid")

        worksheet.row_dimensions[current_row].height = 65


        current_row += 1  # Update current row after adding position details

        # Write player names in the correct position
        top_of_players = current_row
        bottom_of_roster = current_row

        position_keys = ["quarterbacks", "runningbacks", "widereceivers", "tightends", "draftpicks"]

        for j, position in enumerate(position_keys):
            players = roster["players"].get(position, [])
            for k, player in enumerate(players):
                # print(position + ": " + player)
                worksheet.cell(row=current_row + k, column=j + 1, value=player)

            if bottom_of_roster < current_row + k:
                bottom_of_roster = current_row + k

        # Apply alternating gray background to rows of players
        gray = True
        for p in range(top_of_players, bottom_of_roster):
            if gray:
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=p, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                gray = False
            else:
                gray = True

        # # Create a border style for the outer outline
        thin_border = Side(style="thick")
        #
        # # Define the top_of_roster and bottom_of_roster
        # top_of_roster = 10  # Example row where the block starts
        # bottom_of_roster = 20  # Example row where the block ends
        #
        # # The range of columns to be outlined
        first_col = 1
        last_col = 5
        #
        # Add top border to the first row of the block
        for col in range(first_col, last_col + 1):
            worksheet.cell(row=top_of_roster, column=col).border = Border(top=thin_border)
        #
        # # Add bottom border to the last row of the block
        # for col in range(first_col, last_col + 1):
        #     worksheet.cell(row=bottom_of_roster, column=col).border = Border(bottom=thin_border)
        #
        # # Add left border to the leftmost column of the block
        # for row in range(top_of_roster, bottom_of_roster + 1):
        #     worksheet.cell(row=row, column=first_col).border = Border(left=thin_border)
        #
        # # Add right border to the rightmost column of the block
        for row in range(top_of_roster, bottom_of_roster + 1):
            worksheet.cell(row=row, column=last_col).border = Border(right=thin_border)
        for col in range(1, 6):
            worksheet.cell(row=bottom_of_roster, column=col).border = Border(bottom=thin_border)
        worksheet.cell(row=bottom_of_roster, column=5).border = Border(bottom=thin_border, right=thin_border)
        worksheet.cell(row=top_of_roster, column=5).border = Border(top=thin_border, right=thin_border)

        # Update current row for the next roster
        current_row = bottom_of_roster + 4

    # Save the workbook to the specified file
    workbook.save(excel_filename)

