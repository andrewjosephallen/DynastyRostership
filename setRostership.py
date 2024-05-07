import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from getRostership import get_rostership

# Function to set rostership in an Excel file
def set_rostership(workbook, excel_filename, sleeperUsername, playerData):
    if "Rostership" in workbook.sheetnames:
        worksheet = workbook["Rostership"]
    else:
        worksheet = workbook.create_sheet("Rostership")

    # Clear message in a specific cell
    worksheet.cell(row=5, column=5, value="Clearing...")

    # Define some constants for styling
    font_family = "Comfortaa"
    offset = 3
    header_font = Font(bold=True, name=font_family)
    alignment_center = Alignment(horizontal="center")
    fill_light_gray = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")

    # Set column widths
    column_widths = [20, 8, 10, 8, 10, 20]  # Example widths for the six columns
    header_range = range(1 + offset, 1 + offset + len(column_widths))

    for col_idx, width in zip(header_range, column_widths):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Set headers with specific styles
    headers = ["Name", "Pos", "Team", "Num", "%", "Leagues"]
    header_range = range(1 + offset, 1 + offset + len(headers))
    for col_idx, header in zip(header_range, headers):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = alignment_center
        cell.fill = fill_light_gray

    # Retrieve rostership data
    rostership_data = get_rostership(sleeperUsername, playerData)

    # Write data with formatting
    for idx, (player, data) in enumerate(rostership_data.items(), start=2):
        percentage = data["numberOfOccurances"] / data["numberOfLeagues"]* 100
        percentage = f"{percentage:.2f}%"

        row_values = [player, data["position"], data["team"], data["numberOfOccurances"], percentage, data["leagues"]]

        for col_idx, value in zip(header_range, row_values):
            cell = worksheet.cell(row=idx, column=col_idx, value=value)
            cell.font = Font(name=font_family)

            # Apply custom styles (background color for specific columns)
            if col_idx in {1 + offset, 3 + offset, 5 + offset}:
                cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

            # Center alignment for specific columns
            if col_idx in {2 + offset, 3 + offset, 4 + offset, 5 + offset}:
                cell.alignment = alignment_center

    # Save the workbook to the specified file
    workbook.save(excel_filename)